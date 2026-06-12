# -*- coding: utf-8 -*-
import openpyxl, os, re, math
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TMPL = r'E:\Codex Workspace\04_projects\config_templates\telemetry_template.xlsx'
SRC = r'E:\Codex Workspace\04_projects\项目文档\excel版遥测大表草稿\遥测配置表.xlsx'
OUT_DIR = r'E:\Codex Workspace\04_projects\项目文档\优化后excel配置表'
OUT_FILE = os.path.join(OUT_DIR, '遥测配置表_优化后.xlsx')
os.makedirs(OUT_DIR, exist_ok=True)

HF = Font(bold=True, color='FFFFFF', size=10, name='Microsoft YaHei')
HFILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
NF = Font(size=9, name='Microsoft YaHei')
TB = Border(left=Side(style='thin',color='B4C6E7'),right=Side(style='thin',color='B4C6E7'),top=Side(style='thin',color='B4C6E7'),bottom=Side(style='thin',color='B4C6E7'))
EF_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

def sh(ws, r, n):
    for c in range(1, n+1):
        cl = ws.cell(row=r, column=c)
        cl.font = HF; cl.fill = HFILL
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = TB

def sc(ws, r, c, fill=None, center=False):
    cl = ws.cell(row=r, column=c)
    cl.font = NF; cl.border = TB
    cl.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    if fill: cl.fill = fill

# Template header
TMPL_HEADERS = ['序号','遥测代号','参数名称','数据类型','数据域偏移','位偏移(bit)','长度(bit)','字节序','当量','小数位数','单位','初始值(HEX)','下限','上限','说明/枚举取值']

def read_sheet(wb, name):
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None or str(row[0]).strip() == '':
            continue
        r = {}
        r['seq'] = row[0]
        r['param_id'] = str(row[1]).strip() if row[1] else ''
        r['param_name'] = str(row[2]).strip() if row[2] else ''
        r['data_offset'] = row[3]
        r['bit_offset'] = row[4]
        r['bit_length'] = row[5]
        r['data_type'] = str(row[6]).strip() if row[6] else ''
        r['endian'] = str(row[7]).strip() if row[7] else ''
        r['scale'] = row[8]
        r['decimals'] = row[9]
        r['unit'] = str(row[10]).strip() if row[10] else ''
        r['min'] = row[11]
        r['max'] = row[12]
        r['enum_note'] = str(row[13]).strip() if row[13] else ''
        rows.append(r)
    return rows

def fmt_scale(v):
    if v is None or v == '':
        return '1'
    if isinstance(v, (int, float)):
        if v == 1: return '1'
        return f'{v:.10f}'.rstrip('0').rstrip('.')
    return str(v)

def map_type(dt):
    dt = dt.lower().strip()
    m = {'uint8':'uint8','uint16':'uint16','uint24':'uint24','uint32':'uint32',
         'int8':'int8','int16':'int16','int24':'int24','int32':'int32',
         'float32':'float32','float':'float32','enum':'enum','status':'status'}
    return m.get(dt, dt)

def map_endian(e):
    e = e.lower().strip()
    if e in ('big','big_endian','大端'): return 'big_endian'
    if e in ('little','little_endian','小端'): return 'little_endian'
    return e

def parse_enum_values(note):
    if not note: return []
    results = []
    parts = [p.strip() for p in re.split(r'[;；]', note) if p.strip()]
    for p in parts:
        m = re.match(r'(\d+|[0-9A-Fa-f]+)\s*[:：]\s*(.+)', p)
        if m:
            key = m.group(1).strip()
            label = m.group(2).strip()
            results.append((key, label))
    return results

# ===== Read source data =====
src = openpyxl.load_workbook(SRC)
pkg1 = read_sheet(src, '常规包1')
pkg2 = read_sheet(src, '常规包2')
qry1 = read_sheet(src, '查询包1')

# ===== Create new workbook =====
wb = openpyxl.Workbook()
HEADERS = TMPL_HEADERS

def write_sheet(wb, name, data, idx=0):
    if idx == 0:
        ws = wb.active
        ws.title = name
    else:
        ws = wb.create_sheet(name)
    
    for c, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=c, value=h)
    sh(ws, 1, len(HEADERS))
    
    for i, row in enumerate(data, 2):
        dt = map_type(row['data_type'])
        scale = fmt_scale(row['scale'])
        dec = '' if row['decimals'] is None or row['decimals'] == '' else str(int(row['decimals']))
        enum_val = parse_enum_values(row['enum_note'])
        
        bp = '' if row['bit_offset'] is None else str(int(row['bit_offset']))
        bl = '' if row['bit_length'] is None else str(int(row['bit_length']))
        do = '' if row['data_offset'] is None or row['data_offset'] == '' else str(int(row['data_offset']))
        mn = '' if row['min'] is None else str(row['min'])
        mx = '' if row['max'] is None else str(row['max'])
        en = '; '.join(f'{k}:{v}' for k,v in enum_val) if enum_val else (row['enum_note'] or '')
        
        vals = [i-1, row['param_id'], row['param_name'], dt, do, bp, bl,
                map_endian(row['endian']), scale, dec,
                row['unit'] if row['unit'] else '',
                '', mn, mx, en]
        for c, v in enumerate(vals, 1):
            ws.cell(row=i, column=c, value=v)
            sc(ws, i, c, center=(c in (1,4,5,6,7,8,10)))
    
    col_w = [5, 18, 30, 10, 10, 12, 10, 12, 12, 8, 8, 12, 10, 10, 50]
    for i, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(HEADERS))}{len(data)+1}'
    return ws, data

ws1, _ = write_sheet(wb, '常规包1(SlowFrame)', pkg1, 0)

# ===== Sheet 2: 常规包2 =====
ws2, _ = write_sheet(wb, '常规包2(SlowFrame2)', pkg2, 1)

# ===== Sheet 3: 查询包1 =====
ws3, _ = write_sheet(wb, '查询包1(QueryFrame)', qry1, 2)

# ===== Sheet 4: EnumValues (枚举值映射) =====
ws4 = wb.create_sheet('枚举值映射表')
ws4.merge_cells('A1:F1')
nc = ws4.cell(row=1, column=1)
nc.value = '【由"说明/枚举取值"列自动解析生成】'
nc.font = Font(bold=True, color='C00000', size=9, name='Microsoft YaHei')
nc.alignment = Alignment(horizontal='left', vertical='center')
nc.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

h4 = ['所属参数ID','参数名称','数据源(SlowFrame)','Hex值','枚举标签','原始行内容']
for c, h in enumerate(h4, 1): ws4.cell(row=2, column=c, value=h)
sh(ws4, 2, len(h4))

r4 = 3
for label, dataset in [('常规包1', pkg1), ('常规包2', pkg2), ('查询包1', qry1)]:
    for row in dataset:
        ev = parse_enum_values(row['enum_note'])
        if not ev: continue
        for key, label_val in ev:
            vals = [row['param_id'], row['param_name'], label, key, label_val, row['enum_note']]
            for c, v in enumerate(vals, 1):
                ws4.cell(row=r4, column=c, value=v)
                sc(ws4, r4, c, fill=EF_FILL, center=(c in (3,4)))
            r4 += 1

for i, w in enumerate([18, 30, 12, 10, 30, 50], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = 'A3'

# ===== Sheet 5: 帧格式参考 =====
ws5 = wb.create_sheet('帧格式参考(CCSDS)')
# Copy from source
ws_src_fmt = src['帧格式参考(CCSDS)']
h5 = ['层级','区域','字段','位宽(bit)','初始值','取值范围','说明']
for c, h in enumerate(h5, 1): ws5.cell(row=1, column=c, value=h)
sh(ws5, 1, len(h5))

# Enhanced frame format
fd = [
    ['数据包','包主导头','标识符',16,'0x1ACF','固定值','帧头标识'],
    ['','','版本+类型+副导头+APID',16,'0x0520','固定值','高7位设备标识,低4位数据类型'],
    ['','','分组标志',2,'0b11','固定值','单帧数据'],
    ['','','源包序列计数',14,'0','0x00~0x3FFF','14-bit循环累加'],
    ['','','数据域长度',16,'','0x0001~0x07FF','包数据长度减1'],
    ['数据包','包数据域','遥测数据','动态','','','详见SlowFrame/QueryFrame表'],
    ['','包校验域','校验和',16,'','','字节3~N 累加求和取反取低16位'],
]
for i, row in enumerate(fd, 2):
    for c, v in enumerate(row, 1):
        ws5.cell(row=i, column=c, value=v)
        sc(ws5, i, c)
for i, w in enumerate([10,12,35,10,12,18,55], 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = 'A2'

# ===== Save =====
wb.save(OUT_FILE)
t1 = sum(1 for r in pkg1 if parse_enum_values(r['enum_note']))
t2 = sum(1 for r in pkg2 if parse_enum_values(r['enum_note']))
t3 = sum(1 for r in qry1 if parse_enum_values(r['enum_note']))
print(f'DONE: {OUT_FILE}')
print(f'Sheet1 常规包1: {len(pkg1)} 参数')
print(f'Sheet2 常规包2: {len(pkg2)} 参数')
print(f'Sheet3 查询包1: {len(qry1)} 参数')
print(f'Sheet4 枚举值映射: {r4-3} 条 (从 {t1+t2+t3} 个参数解析)')
print(f'Sheet5 帧格式参考')
