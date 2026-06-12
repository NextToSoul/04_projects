# -*- coding: utf-8 -*-
import openpyxl, os, re
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC_FILE = os.path.join("E:\\Codex Workspace\\04_projects", "项目文档", "excel版遥测大表草稿", "立即遥控指令配置表.xlsx")
OUT_DIR = os.path.join("E:\\Codex Workspace\\04_projects", "项目文档", "优化后excel配置表")
OUT_FILE = os.path.join(OUT_DIR, "立即遥控指令配置表_v2.xlsx")
os.makedirs(OUT_DIR, exist_ok=True)

HF = Font(bold=True, color='FFFFFF', size=10, name='Microsoft YaHei')
HFILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
EF = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
TF = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
PF = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
NF = Font(size=9, name='Microsoft YaHei')
TB = Border(left=Side(style='thin',color='B4C6E7'),right=Side(style='thin',color='B4C6E7'),top=Side(style='thin',color='B4C6E7'),bottom=Side(style='thin',color='B4C6E7'))

def norm(v):
    if v is None: return ''
    if isinstance(v, str):
        v = v.replace(' ','').strip()
        return '' if v.upper() in ('','NONE') else v.upper()
    if isinstance(v, (int, float)):
        return str(int(v)) if v == int(v) else str(v)
    return str(v)

def sh(ws, r, n):
    for c in range(1, n+1):
        cl = ws.cell(row=r, column=c)
        cl.font = HF; cl.fill = HFILL
        cl.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cl.border = TB

def sc(ws, r, c, fill=None, center=False):
    cl = ws.cell(row=r, column=c)
    cl.font = NF; cl.border = TB
    cl.alignment = Alignment(horizontal='center' if center else 'left', vertical='center', wrap_text=True)
    if fill: cl.fill = fill

src = openpyxl.load_workbook(SRC_FILE)
ws_src = src['立即遥控指令']

rows = []
for row in ws_src.iter_rows(min_row=2, values_only=True):
    if row[1] is None: continue
    rows.append(dict(
        code_name=str(row[1]).strip(),
        name=str(row[2]).strip() if row[2] else '',
        header=str(row[3]).strip() if row[3] else 'EB 90',
        apid=str(row[4]).strip() if row[4] else '05 20',
        data_len=str(row[5]).strip() if row[5] else '',
        inst_code=str(row[6]).strip().replace(' ','') if row[6] else '',
        param_val=row[7],
        enum_note=str(row[8]).strip() if row[8] else ''
    ))

cg = {}
for row in rows:
    ic = row['inst_code']
    if ic not in cg: cg[ic] = []
    cg[ic].append(row)

enum_codes = set()
for ic, grp in cg.items():
    vals = set(norm(r['param_val']) for r in grp)
    if len(vals) > 1:
        enum_codes.add(ic)

toggle_codes = set()
toggle_notes = {}
for ic, grp in cg.items():
    for r in grp:
        en = r['enum_note']
        if en and ('AAAA' in en.upper().replace(' ','')):
            toggle_codes.add(ic)
            toggle_notes[ic] = en
            break

wb = openpyxl.Workbook()
ws1 = wb.active; ws1.title = '立即遥控指令'

h1 = ['序号','指令代号','控制指令名称','指令码(Hex)','数据域长度(Hex)','参数字节数','默认参数值(Hex)','值类型','下限','上限','转换公式','枚举取值说明']
for c, h in enumerate(h1, 1): ws1.cell(row=1, column=c, value=h)
sh(ws1, 1, len(h1))

r = 2
for idx, row in enumerate(rows, 1):
    ic = row['inst_code']; pv = norm(row['param_val']); en = row['enum_note']
    dl = row['data_len'].replace(' ','')
    pb = len(pv)//2 if pv else 0
    if pv == '': vt = '无参数'; fl = None
    elif ic in enum_codes: vt = '枚举'; fl = EF
    elif ic in toggle_codes: vt = '开关/切换'; fl = TF
    else: vt = '数值参数'; fl = PF
    vals = [idx, row['code_name'], row['name'], ic, dl, pb, pv, vt, '', '', '', en]
    for c, v in enumerate(vals, 1):
        ws1.cell(row=r, column=c, value=v)
        sc(ws1, r, c, fill=fl, center=(c <= 8))
    r += 1

for i, w in enumerate([5,22,30,10,12,8,12,10,6,6,10,50], 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = 'A2'; ws1.auto_filter.ref = f'A1:L{r-1}'

ws2 = wb.create_sheet('枚举值映射表(软件解析)')
ws2.merge_cells('A1:F1')
nc = ws2.cell(row=1, column=1)
nc.value = '【本表由软件自动解析生成，无需手工编辑】'
nc.font = Font(bold=True, color='C00000', size=9, name='Microsoft YaHei')
nc.alignment = Alignment(horizontal='left', vertical='center')
nc.fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')

h2 = ['生成规则', '所属指令码', '参数值(Hex)', '枚举标签', '数据来源', '原始说明']
for c, h in enumerate(h2, 1): ws2.cell(row=2, column=c, value=h)
sh(ws2, 2, len(h2))

r2 = 3
for ic, grp in cg.items():
    if ic not in enum_codes: continue
    for gr in grp:
        pv = norm(gr['param_val'])
        rule = f'同指令码{ic}下参数值不同的行自动分组'
        vals = [rule, ic, pv, gr['name'], '指令名称列', gr['enum_note']]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r2, column=c, value=v)
            sc(ws2, r2, c, fill=EF, center=(c in (2,3)))
        r2 += 1

for ic in toggle_codes:
    en = toggle_notes[ic]
    kv = re.findall(r'([0-9A-F]{2}\s*[0-9A-F]{2}?)\s*[:：]\s*([^;]+)', en)
    for hex_val, label in kv:
        hv = hex_val.replace(' ', '').upper()
        rule = f'从"枚举取值说明"列解析 {hv}'
        vals = [rule, ic, hv, label.strip(), '枚举取值说明列', en]
        for c, v in enumerate(vals, 1):
            ws2.cell(row=r2, column=c, value=v)
            sc(ws2, r2, c, fill=TF, center=(c in (2,3)))
        r2 += 1

for ic, grp in cg.items():
    if ic in enum_codes or ic in toggle_codes: continue
    gr = grp[0]; pv = norm(gr['param_val'])
    if pv == '': continue
    rule = '数值参数，直接使用默认参数值'
    vals = [rule, ic, pv, gr['name'] + ' (数值)', '默认参数值列', gr['enum_note']]
    for c, v in enumerate(vals, 1):
        ws2.cell(row=r2, column=c, value=v)
        sc(ws2, r2, c, fill=PF, center=(c in (2,3)))
    r2 += 1

for i, w in enumerate([40, 10, 10, 30, 18, 50], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A3'

ws3 = wb.create_sheet('帧格式参考(CCSDS)')
h3 = ['层级','区域','字段','位宽(bit)','初始值','取值范围','说明']
for c, h in enumerate(h3, 1): ws3.cell(row=1, column=c, value=h)
sh(ws3, 1, len(h3))

fd = [
    ['数据包','包主导头','标识符',16,'0xEB90','固定值','帧头标识'],
    ['','','版本+类型+副导头+APID',16,'0x0520','固定值','高7位设备标识,低4位数据类型'],
    ['','','分组标志',2,'0b11','固定值','单帧数据'],
    ['','','源包序列计数',14,'0','0x00~0x3FFF','14-bit循环累加'],
    ['','','数据域长度',16,'','0x0001~0x00FF','包数据长度减1'],
    ['数据包','包数据域','指令码',16,'','','见立即遥控指令表'],
    ['','','参数数据','动态','','','视具体指令而定'],
    ['','包校验域','校验和',16,'','','导头(不含标识符)+数据域 单字节累加求和取反取低16bit'],
]
for i, row in enumerate(fd, 2):
    for c, v in enumerate(row, 1):
        ws3.cell(row=i, column=c, value=v); sc(ws3, i, c)
for i, w in enumerate([10,12,35,10,12,18,55], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'

wb.save(OUT_FILE)
print(f'DONE: {OUT_FILE}')

typ = {}
for row in rows:
    ic = row['inst_code']; pv = norm(row['param_val']); en = row['enum_note']
    if pv == '': k = '无参数'
    elif ic in enum_codes: k = '枚举'
    elif ic in toggle_codes: k = '开关/切换'
    else: k = '数值参数'
    typ[k] = typ.get(k, 0) + 1
print(f'Sheet1: {len(rows)} 条指令')
for k, v in sorted(typ.items(), key=lambda x: -x[1]): print(f'  {k}: {v}')
print(f'Sheet2: {r2-3} 条映射')
print(f'Sheet3: 帧格式参考')
