import openpyxl

def load_excel_config(filepath, sheet_name):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows

def normalize_hex(val):
    if val is None:
        return ''
    if isinstance(val, str):
        v = val.replace(' ', '').strip()
        return v.upper() if v else ''
    if isinstance(val, (int, float)):
        return str(int(val))
    return str(val)
