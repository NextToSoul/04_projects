import openpyxl
from src.config.loader import load_excel_config

class InjectTable:
    def __init__(self, filepath):
        self.filepath = filepath
        self.params = []
        self._load()
    
    def _load(self):
        wb = openpyxl.load_workbook(self.filepath, data_only=True)
        for sname in wb.sheetnames:
            if '固定地址参数注入' in sname and '帧格式' not in sname:
                self.sheet_name = sname
                break
        else:
            raise ValueError('注入参数Sheet未找到: ' + str(wb.sheetnames))
        self.params = load_excel_config(self.filepath, self.sheet_name)
    
    def get_params(self):
        return self.params
    
    def lookup(self, param_id):
        for p in self.params:
            if str(p.get('参数ID', '') or '').strip() == param_id:
                return p
        raise KeyError('参数 ' + param_id + ' 未找到')
    
    @property
    def instruction_code(self):
        if '注入1' in self.sheet_name:
            return '03FF'
        return '03FE'
    
    def total_bit_length(self):
        key = '位长度' + chr(10) + 'bit_length'
        return sum(int(p.get(key, 0) or 0) for p in self.params)
